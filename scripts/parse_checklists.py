#!/usr/bin/env python3
"""Excel checklist dosyalarını JSON'a dönüştürür. Bir kerelik ve güncellemelerde çalışır."""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl gerekli. Kurulum: pip install openpyxl")


SEO_SHEET_MAP = {
    'seo-ecommerce': 'E-commerce Checklist',
    'seo-local': 'Local Checklist',
    'seo-enterprise': 'Enterprise Checklist',
    'seo-international': 'International Checklist',
    'seo-blog': 'Blog Checklist',
}
GEO_SHEET_MAP = {
    'geo-eticaret': 'E-Ticaret',
    'geo-hizmet': 'Hizmet',
}


def parse_seo_sheet(ws):
    items = []
    headers = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(row):
            continue
        if headers is None and row[0] == 'Type' and row[1] == 'Task ':
            headers = row
            continue
        if headers is None:
            continue
        type_val, task, questions, guides, est_time = row[0], row[1], row[2], row[3], row[4]
        if not (type_val or task):
            continue
        items.append({
            'type': (type_val or '').strip() if type_val else '',
            'task': (task or '').strip() if task else '',
            'questions': (questions or '').strip() if questions else '',
            'guides': (guides or '').strip() if guides else '',
            'estimated_time': (str(est_time).strip() if est_time else ''),
        })
    return items


def parse_geo_sheet(ws):
    items = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(row):
            continue
        if row[0] and isinstance(row[0], str) and row[0].startswith('FAZ'):
            # Phase heading row, skip
            continue
        if row[0] == '☐' and row[1] == 'Faz':
            # Header row, skip (don't add as item)
            continue
        if row[0] == '☐' and row[1] and isinstance(row[1], str) and row[1].startswith('Faz '):
            # Data row
            items.append({
                'phase': str(row[1]).strip(),
                'task_type': (row[2] or '').strip() if row[2] else '',
                'channel': (row[3] or '').strip() if row[3] else '',
                'action': (row[4] or '').strip() if row[4] else '',
                'detail': (row[5] or '').strip() if row[5] else '',
                'tool': (row[6] or '').strip() if row[6] else '',
                'priority': (row[7] or '').strip() if row[7] else '',
                'owner': (row[8] or '').strip() if row[8] else '',
            })
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--seo', required=True, help='SEO checklist xlsx yolu')
    ap.add_argument('--geo', required=True, help='GEO checklist xlsx yolu')
    ap.add_argument('--out', required=True, help='Çıktı dizini')
    args = ap.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    import warnings
    warnings.filterwarnings('ignore')

    seo_wb = openpyxl.load_workbook(args.seo, data_only=True)
    for json_name, sheet_name in SEO_SHEET_MAP.items():
        if sheet_name not in seo_wb.sheetnames:
            print(f"skip: {sheet_name} yok", file=sys.stderr)
            continue
        items = parse_seo_sheet(seo_wb[sheet_name])
        (out_dir / f'{json_name}.json').write_text(
            json.dumps({'version': '1.0.0', 'source_sheet': sheet_name, 'items': items},
                       ensure_ascii=False, indent=2, sort_keys=True)
        )
        print(f"✓ {json_name}.json ({len(items)} item)")

    geo_wb = openpyxl.load_workbook(args.geo, data_only=True)
    for json_name, sheet_name in GEO_SHEET_MAP.items():
        if sheet_name not in geo_wb.sheetnames:
            print(f"skip: {sheet_name} yok", file=sys.stderr)
            continue
        items = parse_geo_sheet(geo_wb[sheet_name])
        (out_dir / f'{json_name}.json').write_text(
            json.dumps({'version': '1.0.0', 'source_sheet': sheet_name, 'items': items},
                       ensure_ascii=False, indent=2, sort_keys=True)
        )
        print(f"✓ {json_name}.json ({len(items)} item)")


if __name__ == '__main__':
    main()
